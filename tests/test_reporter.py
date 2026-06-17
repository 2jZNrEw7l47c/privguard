# tests/test_reporter.py
import json
import pytest
from datetime import date
from pathlib import Path

from privguard.db import init_db, DB_PATH
from privguard.reporter import generate_report

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _insert_finding(conn, **kwargs):
    defaults = dict(
        user_display_name="John Smith",
        source="broker",
        site_id="testsite",
        site_name="Test Site",
        data_found=None,
        status="found",
        opt_out_url="https://example.com/optout",
        manual_instructions=None,
        screenshot_path=None,
        last_checked="2026-06-15 20:00:00",
        last_submitted=None,
        notes=None,
    )
    defaults.update(kwargs)
    conn.execute(
        """
        INSERT INTO findings
          (user_display_name, source, site_id, site_name, data_found, status,
           opt_out_url, manual_instructions, screenshot_path,
           last_checked, last_submitted, notes)
        VALUES
          (:user_display_name, :source, :site_id, :site_name, :data_found, :status,
           :opt_out_url, :manual_instructions, :screenshot_path,
           :last_checked, :last_submitted, :notes)
        """,
        defaults,
    )


def _insert_breach(conn, **kwargs):
    defaults = dict(
        user_display_name="John Smith",
        email="john@gmail.com",
        breach_name="Adobe",
        breach_date="2013-10-04",
        exposed_fields=json.dumps(["Email addresses", "Passwords"]),
        hibp_url="https://haveibeenpwned.com/account/john@gmail.com",
        added_at="2026-06-15 20:00:00",
    )
    defaults.update(kwargs)
    conn.execute(
        """
        INSERT INTO breaches
          (user_display_name, email, breach_name, breach_date,
           exposed_fields, hibp_url, added_at)
        VALUES
          (:user_display_name, :email, :breach_name, :breach_date,
           :exposed_fields, :hibp_url, :added_at)
        """,
        defaults,
    )


def _profile(name="John Smith"):
    return {"display_name": name}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def db(tmp_path):
    """Return an open sqlite3 connection to a fresh test DB."""
    import sqlite3

    db_file = tmp_path / "test.db"
    conn = sqlite3.connect(db_file)
    init_db(db_path=db_file)
    yield conn, db_file
    conn.close()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestGenerateReportFile:
    def test_creates_xlsx_file(self, tmp_path, db):
        conn, db_file = db
        _insert_finding(conn)
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)

        assert out.suffix == ".xlsx"
        assert out.exists()

    def test_filename_contains_display_name_and_date(self, tmp_path, db):
        conn, db_file = db
        conn.commit()

        out = generate_report(_profile("Jane Doe"), tmp_path, db_path=db_file)
        today = date.today().isoformat()

        assert "Jane_Doe" in out.name
        assert today in out.name

    def test_filename_spaces_replaced_with_underscores(self, tmp_path, db):
        conn, db_file = db
        conn.commit()

        out = generate_report(_profile("Alice Bob Carol"), tmp_path, db_path=db_file)

        assert "Alice_Bob_Carol" in out.name
        assert " " not in out.name


class TestSheetNames:
    def test_has_exactly_five_sheets(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)

        assert len(wb.sheetnames) == 5

    def test_sheet_names_correct(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)

        assert wb.sheetnames == [
            "Summary",
            "Data Brokers",
            "Breaches (HIBP)",
            "Social Platforms",
            "Search Engine Removals",
        ]


class TestDataBrokersSheet:
    def test_headers_correct(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "Site Name",
            "Status",
            "Data Found",
            "Opt-Out URL",
            "Date Submitted",
            "Screenshot Saved",
            "Manual Instructions",
        ]

    def test_status_cell_red_fill_for_found(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        status_cell = ws.cell(row=2, column=2)
        assert status_cell.fill.fgColor.rgb == "FFFF4444"

    def test_status_cell_green_fill_for_submitted(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", status="submitted")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        status_cell = ws.cell(row=2, column=2)
        assert status_cell.fill.fgColor.rgb == "FF44AA44"

    def test_status_cell_orange_fill_for_manual_required(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", status="manual_required")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        status_cell = ws.cell(row=2, column=2)
        assert status_cell.fill.fgColor.rgb == "FFFF8800"

    def test_status_cell_yellow_fill_for_pending_verification(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", status="pending_verification")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        status_cell = ws.cell(row=2, column=2)
        assert status_cell.fill.fgColor.rgb == "FFFFDD00"

    def test_status_cell_grey_fill_for_not_found(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", status="not_found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        status_cell = ws.cell(row=2, column=2)
        assert status_cell.fill.fgColor.rgb == "FFAAAAAA"

    def test_only_broker_findings_appear(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="broker", site_name="BrokerSite")
        _insert_finding(conn, source="social", site_name="SocialSite", site_id="social1")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Data Brokers"]

        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "BrokerSite" in names
        assert "SocialSite" not in names


class TestSummarySheet:
    def _get_summary_value(self, ws, field_name):
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == field_name:
                return row[1]
        return None

    def test_risk_score_high_when_exposures_above_15(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        for i in range(16):
            _insert_finding(conn, site_id=f"site{i}", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Exposure Risk Score") == "HIGH"

    def test_risk_score_medium_when_exposures_between_5_and_15(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        for i in range(10):
            _insert_finding(conn, site_id=f"site{i}", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Exposure Risk Score") == "MEDIUM"

    def test_risk_score_low_when_exposures_below_5(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        for i in range(3):
            _insert_finding(conn, site_id=f"site{i}", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Exposure Risk Score") == "LOW"

    def test_risk_score_medium_at_boundary_5(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        for i in range(5):
            _insert_finding(conn, site_id=f"site{i}", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Exposure Risk Score") == "MEDIUM"

    def test_risk_score_high_at_boundary_16(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        for i in range(16):
            _insert_finding(conn, site_id=f"site{i}", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Exposure Risk Score") == "HIGH"

    def test_summary_counts_auto_submitted(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, site_id="s1", status="submitted")
        _insert_finding(conn, site_id="s2", status="submitted")
        _insert_finding(conn, site_id="s3", status="found")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Auto-Submitted") == 2

    def test_summary_counts_breaches(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_breach(conn, breach_name="Adobe")
        _insert_breach(conn, breach_name="LinkedIn")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Breaches Found") == 2

    def test_summary_profile_name(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        conn.commit()

        out = generate_report(_profile("Alice Smith"), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Summary"]

        assert self._get_summary_value(ws, "Profile Name") == "Alice Smith"


class TestBreachesSheet:
    def test_headers_correct(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_breach(conn)
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Breaches (HIBP)"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == [
            "Email",
            "Breach Name",
            "Breach Date",
            "Exposed Data Types",
            "HIBP Link",
            "Recommended Action",
        ]

    def test_exposed_fields_parsed_from_json(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_breach(conn, exposed_fields=json.dumps(["Email addresses", "Passwords"]))
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Breaches (HIBP)"]

        exposed = ws.cell(row=2, column=4).value
        assert exposed == "Email addresses, Passwords"

    def test_recommended_action_text(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_breach(conn)
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Breaches (HIBP)"]

        action = ws.cell(row=2, column=6).value
        assert action == (
            "Change password for this service and any reused passwords. Enable 2FA."
        )


class TestSocialPlatformsSheet:
    def test_headers_correct(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Social Platforms"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == [
            "Platform",
            "Profile URL",
            "Public Fields Exposed",
            "Recommended Action",
        ]

    def test_only_social_findings_appear(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="social", site_name="Facebook", site_id="facebook1")
        _insert_finding(conn, source="broker", site_name="Whitepages", site_id="whitepages1")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Social Platforms"]

        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Facebook" in names
        assert "Whitepages" not in names


class TestSearchEngineRemovalsSheet:
    def test_headers_correct(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Search Engine Removals"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == [
            "Engine",
            "Status",
            "Removal Request URL",
            "Date Submitted",
        ]

    def test_only_search_engine_findings_appear(self, tmp_path, db):
        import openpyxl

        conn, db_file = db
        _insert_finding(conn, source="search_engine", site_name="Google", site_id="google1")
        _insert_finding(conn, source="broker", site_name="Whitepages", site_id="whitepages1")
        conn.commit()

        out = generate_report(_profile(), tmp_path, db_path=db_file)
        wb = openpyxl.load_workbook(out)
        ws = wb["Search Engine Removals"]

        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "Google" in names
        assert "Whitepages" not in names
