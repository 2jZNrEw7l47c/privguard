"""
Excel report generator for PrivGuard.

Public API:
    generate_report(profile, output_dir, db_path=DB_PATH) -> Path
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from privguard.db import DB_PATH, get_breaches, get_findings

_HEADER_FILL = PatternFill("solid", fgColor="FF2B4099")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_STATUS_FILLS: dict[str, PatternFill] = {
    "found": PatternFill("solid", fgColor="FFFF4444"),
    "manual_required": PatternFill("solid", fgColor="FFFF8800"),
    "pending_verification": PatternFill("solid", fgColor="FFFFDD00"),
    "submitted": PatternFill("solid", fgColor="FF44AA44"),
    "cleared": PatternFill("solid", fgColor="FF44AA44"),
    "not_found": PatternFill("solid", fgColor="FFAAAAAA"),
}

_MAX_COL_WIDTH = 60


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, _MAX_COL_WIDTH)


def _exposure_risk_score(exposure_count: int) -> str:
    if exposure_count > 15:
        return "HIGH"
    if exposure_count >= 5:
        return "MEDIUM"
    return "LOW"


def _build_summary(ws, profile: dict, findings: list[dict], breaches: list[dict]) -> None:
    ws.title = "Summary"

    exposures = [f for f in findings if f["status"] != "not_found"]
    auto_submitted = [f for f in findings if f["status"] == "submitted"]
    pending = [f for f in findings if f["status"] == "pending_verification"]
    manual_req = [f for f in findings if f["status"] == "manual_required"]

    rows = [
        ("Field", "Value"),
        ("Profile Name", profile["display_name"]),
        ("Scan Date", date.today().isoformat()),
        ("Total Sites Checked", len(findings)),
        ("Exposures Found", len(exposures)),
        ("Auto-Submitted", len(auto_submitted)),
        ("Pending Email Verification", len(pending)),
        ("Manual Action Required", len(manual_req)),
        ("Breaches Found", len(breaches)),
        ("Exposure Risk Score", _exposure_risk_score(len(exposures))),
    ]

    for row_idx, (field, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=field)
        ws.cell(row=row_idx, column=2, value=value)

    _style_header_row(ws, num_cols=2)
    _autosize_columns(ws)


def _build_data_brokers(ws, findings: list[dict]) -> None:
    ws.title = "Data Brokers"

    headers = [
        "Site Name",
        "Status",
        "Data Found",
        "Opt-Out URL",
        "Date Submitted",
        "Screenshot Saved",
        "Manual Instructions",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, num_cols=len(headers))

    broker_findings = [f for f in findings if f["source"] == "brokers"]
    for row_idx, finding in enumerate(broker_findings, start=2):
        ws.cell(row=row_idx, column=1, value=finding["site_name"])
        status_cell = ws.cell(row=row_idx, column=2, value=finding["status"])
        ws.cell(row=row_idx, column=3, value=finding["data_found"])
        ws.cell(row=row_idx, column=4, value=finding["opt_out_url"])
        ws.cell(row=row_idx, column=5, value=finding["last_submitted"])
        ws.cell(
            row=row_idx,
            column=6,
            value="Yes" if finding["screenshot_path"] else "No",
        )
        ws.cell(row=row_idx, column=7, value=finding["manual_instructions"])

        status_fill = _STATUS_FILLS.get(finding["status"])
        if status_fill:
            status_cell.fill = status_fill

    _autosize_columns(ws)


def _build_breaches(ws, breaches: list[dict]) -> None:
    ws.title = "Breaches (HIBP)"

    headers = [
        "Email",
        "Breach Name",
        "Breach Date",
        "Exposed Data Types",
        "HIBP Link",
        "Recommended Action",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, num_cols=len(headers))

    recommended = (
        "Change password for this service and any reused passwords. Enable 2FA."
    )

    for row_idx, breach in enumerate(breaches, start=2):
        try:
            fields = json.loads(breach["exposed_fields"] or "[]")
        except (json.JSONDecodeError, TypeError):
            fields = []

        ws.cell(row=row_idx, column=1, value=breach["email"])
        ws.cell(row=row_idx, column=2, value=breach["breach_name"])
        ws.cell(row=row_idx, column=3, value=breach["breach_date"])
        ws.cell(row=row_idx, column=4, value=", ".join(fields))
        ws.cell(row=row_idx, column=5, value=breach["hibp_url"])
        ws.cell(row=row_idx, column=6, value=recommended)

    _autosize_columns(ws)


def _build_social_platforms(ws, findings: list[dict]) -> None:
    ws.title = "Social Platforms"

    headers = [
        "Platform",
        "Profile URL",
        "Public Fields Exposed",
        "Recommended Action",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, num_cols=len(headers))

    social_findings = [f for f in findings if f["source"] == "social"]
    for row_idx, finding in enumerate(social_findings, start=2):
        ws.cell(row=row_idx, column=1, value=finding["site_name"])
        ws.cell(row=row_idx, column=2, value=finding["opt_out_url"])
        ws.cell(row=row_idx, column=3, value=finding["data_found"])
        ws.cell(row=row_idx, column=4, value=finding["notes"])

    _autosize_columns(ws)


def _build_search_engine_removals(ws, findings: list[dict]) -> None:
    ws.title = "Search Engine Removals"

    headers = [
        "Engine",
        "Status",
        "Removal Request URL",
        "Date Submitted",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, num_cols=len(headers))

    se_findings = [f for f in findings if f["source"] == "search_engines"]
    for row_idx, finding in enumerate(se_findings, start=2):
        ws.cell(row=row_idx, column=1, value=finding["site_name"])
        ws.cell(row=row_idx, column=2, value=finding["status"])
        ws.cell(row=row_idx, column=3, value=finding["opt_out_url"])
        ws.cell(row=row_idx, column=4, value=finding["last_submitted"])

    _autosize_columns(ws)


def generate_report(
    profile: dict,
    output_dir: Path,
    db_path: Path = DB_PATH,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    display_name = profile["display_name"]
    safe_name = display_name.replace(" ", "_")
    today = date.today().isoformat()
    filename = f"PrivGuard_Report_{safe_name}_{today}.xlsx"
    out_path = output_dir / filename

    findings = get_findings(user_display_name=display_name, db_path=db_path)
    breaches = get_breaches(user_display_name=display_name, db_path=db_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_summary(wb.create_sheet("Summary"), profile, findings, breaches)
    _build_data_brokers(wb.create_sheet("Data Brokers"), findings)
    _build_breaches(wb.create_sheet("Breaches (HIBP)"), breaches)
    _build_social_platforms(wb.create_sheet("Social Platforms"), findings)
    _build_search_engine_removals(wb.create_sheet("Search Engine Removals"), findings)

    wb.save(out_path)
    return out_path
